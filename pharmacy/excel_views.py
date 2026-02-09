import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.shortcuts import redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from datetime import datetime
from decimal import Decimal, InvalidOperation
from .models import Categorie, Fournisseur, Produit, Client, Vente, LigneVente


def admin_required(view_func):
    from functools import wraps

    @wraps(view_func)
    @login_required
    def wrapper(request, *args, **kwargs):
        if not request.user.is_admin:
            messages.error(request, "Accès réservé à l'administrateur.")
            return redirect('dashboard')
        return view_func(request, *args, **kwargs)
    return wrapper


# ===================== STYLES =====================

HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
HEADER_FILL = PatternFill(start_color='1a2332', end_color='1a2332', fill_type='solid')
HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)


def style_header(ws, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
            cell.border = THIN_BORDER
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


def make_response(wb, filename):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# ===================== EXPORTS =====================

@admin_required
def export_categories(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Catégories'
    ws.append(['Code', 'Désignation'])
    style_header(ws, 2)
    for c in Categorie.objects.all():
        ws.append([c.code_categorie, c.designation])
    auto_width(ws)
    return make_response(wb, 'categories.xlsx')


@admin_required
def export_fournisseurs(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Fournisseurs'
    ws.append(['Code', 'Désignation', 'Marge Bénéficiaire (%)'])
    style_header(ws, 3)
    for f in Fournisseur.objects.all():
        ws.append([f.code_fournisseur, f.designation, float(f.marge_beneficiaire)])
    auto_width(ws)
    return make_response(wb, 'fournisseurs.xlsx')


@admin_required
def export_produits(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Produits'
    ws.append(['Code', 'Désignation', 'Prix Achat', 'Qté Initiale', 'Qté Stock',
               'Qté Alerte', 'Jours Alerte Exp.', 'Fournisseur', 'Date Expiration', 'Catégorie', 'Prix Vente'])
    style_header(ws, 11)
    for p in Produit.objects.select_related('fournisseur', 'categorie').all():
        ws.append([
            p.code_produit, p.designation, float(p.prix_achat),
            p.quantite_initiale, p.quantite_stock, p.quantite_alerte,
            p.jours_alerte_expiration, p.fournisseur.designation,
            p.date_expiration.strftime('%d/%m/%Y') if p.date_expiration else '',
            p.categorie.designation, float(p.prix_vente)
        ])
    auto_width(ws)
    return make_response(wb, 'produits.xlsx')


@admin_required
def export_clients(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Clients'
    ws.append(['Code', 'Nom', 'Téléphone', 'Adresse'])
    style_header(ws, 4)
    for c in Client.objects.all():
        ws.append([c.code_client, c.nom, c.telephone, c.adresse])
    auto_width(ws)
    return make_response(wb, 'clients.xlsx')


@admin_required
def export_ventes(request):
    wb = openpyxl.Workbook()

    # Feuille 1 : Ventes
    ws = wb.active
    ws.title = 'Ventes'
    ws.append(['Code', 'Date', 'Client', 'Type', 'Vendeur', 'Montant Total'])
    style_header(ws, 6)
    for v in Vente.objects.select_related('client', 'vendeur').all():
        ws.append([
            v.code_vente,
            v.date_vente.strftime('%d/%m/%Y %H:%M'),
            str(v.client) if v.client else 'Client Anonyme',
            v.get_type_vente_display(),
            v.vendeur.get_full_name() or v.vendeur.username,
            float(v.montant_total)
        ])
    auto_width(ws)

    # Feuille 2 : Lignes de vente
    ws2 = wb.create_sheet('Lignes de Vente')
    ws2.append(['Code Vente', 'Produit', 'Quantité', 'Prix Unitaire', 'Montant Ligne'])
    style_header(ws2, 5)
    for l in LigneVente.objects.select_related('vente', 'produit').all():
        ws2.append([
            l.vente.code_vente, l.produit.designation,
            l.quantite, float(l.prix_unitaire), float(l.montant_ligne)
        ])
    auto_width(ws2)

    return make_response(wb, 'ventes.xlsx')


# ===================== IMPORTS =====================

@admin_required
def import_categories(request):
    if request.method == 'POST' and request.FILES.get('fichier_excel'):
        try:
            wb = openpyxl.load_workbook(request.FILES['fichier_excel'])
            ws = wb.active
            count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[1]:
                    Categorie.objects.get_or_create(designation=str(row[1]).strip())
                    count += 1
            messages.success(request, f"{count} catégorie(s) traitée(s).")
        except Exception as e:
            messages.error(request, f"Erreur lors de l'import : {e}")
    return redirect('categorie_list')


@admin_required
def import_fournisseurs(request):
    if request.method == 'POST' and request.FILES.get('fichier_excel'):
        try:
            wb = openpyxl.load_workbook(request.FILES['fichier_excel'])
            ws = wb.active
            count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[1] and row[2] is not None:
                    Fournisseur.objects.get_or_create(
                        designation=str(row[1]).strip(),
                        defaults={'marge_beneficiaire': Decimal(str(row[2]))}
                    )
                    count += 1
            messages.success(request, f"{count} fournisseur(s) traité(s).")
        except Exception as e:
            messages.error(request, f"Erreur lors de l'import : {e}")
    return redirect('fournisseur_list')


@admin_required
def import_produits(request):
    if request.method == 'POST' and request.FILES.get('fichier_excel'):
        try:
            wb = openpyxl.load_workbook(request.FILES['fichier_excel'])
            ws = wb.active
            count = 0
            errors = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                try:
                    designation = str(row[1]).strip() if row[1] else None
                    if not designation:
                        continue
                    prix_achat = Decimal(str(row[2]))
                    qte_initiale = int(row[3]) if row[3] else 0
                    qte_stock = int(row[4]) if row[4] else qte_initiale
                    qte_alerte = int(row[5]) if row[5] else 10
                    jours_alerte = int(row[6]) if row[6] else 30
                    fournisseur_nom = str(row[7]).strip() if row[7] else None
                    date_exp_raw = row[8]
                    categorie_nom = str(row[9]).strip() if row[9] else None

                    if not fournisseur_nom or not categorie_nom:
                        errors += 1
                        continue

                    fournisseur, _ = Fournisseur.objects.get_or_create(
                        designation=fournisseur_nom, defaults={'marge_beneficiaire': 0})
                    categorie, _ = Categorie.objects.get_or_create(designation=categorie_nom)

                    if isinstance(date_exp_raw, datetime):
                        date_exp = date_exp_raw.date()
                    elif isinstance(date_exp_raw, str):
                        for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
                            try:
                                date_exp = datetime.strptime(date_exp_raw.strip(), fmt).date()
                                break
                            except ValueError:
                                continue
                        else:
                            errors += 1
                            continue
                    else:
                        date_exp = date_exp_raw

                    Produit.objects.get_or_create(
                        designation=designation,
                        defaults={
                            'prix_achat': prix_achat,
                            'quantite_initiale': qte_initiale,
                            'quantite_stock': qte_stock,
                            'quantite_alerte': qte_alerte,
                            'jours_alerte_expiration': jours_alerte,
                            'fournisseur': fournisseur,
                            'date_expiration': date_exp,
                            'categorie': categorie,
                        }
                    )
                    count += 1
                except (ValueError, InvalidOperation, TypeError):
                    errors += 1
                    continue
            msg = f"{count} produit(s) traité(s)."
            if errors:
                msg += f" {errors} ligne(s) ignorée(s)."
            messages.success(request, msg)
        except Exception as e:
            messages.error(request, f"Erreur lors de l'import : {e}")
    return redirect('produit_list')


@admin_required
def import_clients(request):
    if request.method == 'POST' and request.FILES.get('fichier_excel'):
        try:
            wb = openpyxl.load_workbook(request.FILES['fichier_excel'])
            ws = wb.active
            count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                nom = str(row[1]).strip() if row[1] else None
                if not nom:
                    continue
                Client.objects.get_or_create(
                    nom=nom,
                    defaults={
                        'telephone': str(row[2]).strip() if row[2] else '',
                        'adresse': str(row[3]).strip() if row[3] else '',
                    }
                )
                count += 1
            messages.success(request, f"{count} client(s) traité(s).")
        except Exception as e:
            messages.error(request, f"Erreur lors de l'import : {e}")
    return redirect('client_list')
